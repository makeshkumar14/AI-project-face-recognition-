"""
AI-Powered Face Recognition Attendance System
Faculty-Only Flask Backend
"""

from datetime import datetime
import os
import glob
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, send_from_directory
from flask_cors import CORS
import threading

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'attend-secret-2024')
CORS(app, resources={r"/api/*": {"origins": "*"}})

# ─── Database Init ────────────────────────────────────────────────────────────
from models import init_db, seed_sample_data
with app.app_context():
    init_db()
    seed_sample_data()

# ─── Pre-load models in background ───────────────────────────────────────────
def preload_models():
    try:
        from advanced_face_recognition import get_recognizer
        get_recognizer()
        print("Advanced Face Recognition models pre-loaded successfully!")
    except Exception as e:
        print(f"Could not pre-load advanced models: {e}")

threading.Thread(target=preload_models, daemon=True).start()

# ─── Blueprints ───────────────────────────────────────────────────────────────
from routes.auth import auth_bp
from routes.attendance import attendance_bp
app.register_blueprint(auth_bp)
app.register_blueprint(attendance_bp)

# ─── Helper: require faculty login ───────────────────────────────────────────
def require_faculty(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('user_type') != 'faculty':
            return redirect(url_for('faculty_login'))
        return f(*args, **kwargs)
    return decorated

# ─── Auth Routes ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if session.get('user_type') == 'faculty':
        return redirect(url_for('faculty_profile'))
    return redirect(url_for('faculty_login'))


@app.route('/faculty_login', methods=['GET', 'POST'])
def faculty_login():
    """Faculty login + register page."""
    message = request.args.get('message')
    message_type = request.args.get('message_type', 'error')

    if request.method == 'POST':
        action = request.form.get('action', 'login')

        if action == 'register':
            name = request.form.get('name', '').strip()
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '')
            department = request.form.get('department', 'Computer Science').strip()

            if not name or not email or not password:
                return render_template('faculty_login.html', message='All fields are required.', message_type='error', show_register=True)

            from models import add_faculty
            if add_faculty(name, email, password, department):
                return redirect(url_for('faculty_login', message='Registration successful! Please log in.', message_type='success'))
            else:
                return render_template('faculty_login.html', message='Email already registered.', message_type='error', show_register=True)

        else:  # login
            email = request.form.get('faculty_id', '').strip()
            password = request.form.get('password', '')

            from models import verify_faculty_password
            faculty = verify_faculty_password(email, password)

            if faculty:
                session['user_type'] = 'faculty'
                session['user_id'] = faculty['id']
                session['user_name'] = faculty['name']
                session['user_email'] = faculty['email']
                session['user_dept'] = faculty.get('department', '')
                return redirect(url_for('faculty_profile'))
            else:
                return render_template('faculty_login.html',
                                       message='Invalid email or password.',
                                       message_type='error',
                                       show_register_option=True)

    return render_template('faculty_login.html', message=message, message_type=message_type)


@app.route('/faculty_register', methods=['POST'])
def faculty_register():
    """Standalone faculty registration (also handled in faculty_login)."""
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')
    confirm_password = request.form.get('confirm_password', '')
    department = request.form.get('department', 'Computer Science').strip()

    if not name or not email or not password:
        return render_template('faculty_login.html', message='All fields required.', message_type='error', show_register=True)
    if password != confirm_password:
        return render_template('faculty_login.html', message='Passwords do not match.', message_type='error', show_register=True)
    if len(password) < 4:
        return render_template('faculty_login.html', message='Password must be at least 4 characters.', message_type='error', show_register=True)

    from models import add_faculty
    if add_faculty(name, email, password, department):
        return redirect(url_for('faculty_login', message='Registration successful! Please sign in.', message_type='success'))
    else:
        return render_template('faculty_login.html', message='Email already registered.', message_type='error', show_register=True)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('faculty_login'))


# ─── Faculty Profile Dashboard ───────────────────────────────────────────────
@app.route('/faculty_profile')
@require_faculty
def faculty_profile():
    """Dedicated Faculty Profile Dashboard - replaces old select_subject flow."""
    return render_template('faculty_profile.html')


@app.route('/edit_profile', methods=['GET', 'POST'])
@require_faculty
def edit_profile():
    """Edit faculty personal details."""
    from models import update_faculty
    faculty_id = session.get('user_id')
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        department = request.form.get('department', '').strip()
        
        if not name or not email:
            return render_template('edit_profile.html', 
                                 name=name, email=email, department=department,
                                 error="Name and Email are required.")
        
        if update_faculty(faculty_id, name, email, department):
            # Update session
            session['user_name'] = name
            session['user_email'] = email
            session['user_dept'] = department
            return redirect(url_for('faculty_profile'))
        else:
            return render_template('edit_profile.html', 
                                 name=name, email=email, department=department,
                                 error="Could not update profile. Email might already be in use.")

    return render_template('edit_profile.html',
                         name=session.get('user_name'),
                         email=session.get('user_email'),
                         department=session.get('user_dept'))


# ─── Bridge: Profile → Webcam Attendance Page ────────────────────────────────
@app.route('/begin_attendance')
@require_faculty
def begin_attendance():
    """
    Called from faculty_profile form submission.
    Saves selection to session, starts attendance session, redirects to webcam dashboard.
    """
    subject = request.args.get('subject', '').strip()
    department = request.args.get('department', '').strip()
    section = request.args.get('section', '').strip()
    period = request.args.get('period', '1').strip()

    if not subject or not section:
        return redirect(url_for('faculty_profile'))

    # Store in session for the webcam dashboard
    session['current_subject'] = subject
    session['current_department'] = department
    session['current_section'] = section
    session['current_period'] = period

    # Start the attendance logic session (pre-initialises before webcam page loads)
    from attendance_logic import start_attendance_session
    faculty_id = session.get('user_id')
    # Force=True ensures any old session metadata is wiped
    start_attendance_session(subject, section, department, period, faculty_id=faculty_id, force=True)

    return redirect(url_for('faculty_dashboard'))


@app.route('/faculty_dashboard')
@require_faculty
def faculty_dashboard():
    """Webcam attendance tracking page (unchanged)."""
    return render_template('faculty_dashboard.html')


@app.route('/faculty_reports')
@require_faculty
def faculty_reports():
    """Attendance reports page."""
    return render_template('faculty_reports.html')


# ─── Faculty Dashboard API Endpoints ─────────────────────────────────────────
@app.route('/api/faculty/profile')
@require_faculty
def api_faculty_profile():
    """Return logged-in faculty profile info."""
    return jsonify({
        'success': True,
        'name': session.get('user_name', ''),
        'email': session.get('user_email', ''),
        'department': session.get('user_dept', ''),
    })


@app.route('/api/faculty/stats')
@require_faculty
def api_faculty_stats():
    """Return stat card values for the profile dashboard."""
    from models import get_faculty_stats
    faculty_id = session.get('user_id')
    stats = get_faculty_stats(faculty_id)

    # Compute avg attendance %
    total_sessions = stats.get('total_sessions') or 0
    total_present  = stats.get('total_present') or 0
    unique_subjects = stats.get('unique_subjects') or 0

    # We need total students marked to calc avg properly; use a quick heuristic
    from models import get_db_connection
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT COUNT(*) FROM attendance WHERE faculty_id = %s", (faculty_id,)
        )
        total_records = cursor.fetchone()[0] or 0
        cursor.close()
        conn.close()
    except Exception:
        total_records = 0

    avg = round((total_present / total_records * 100), 1) if total_records > 0 else 0

    return jsonify({
        'success': True,
        'total_sessions': total_sessions,
        'total_students_marked': total_present,
        'avg_attendance': avg,
        'unique_subjects': unique_subjects,
    })


@app.route('/api/faculty/history')
@require_faculty
def api_faculty_history():
    """Return past attendance sessions for the logged-in faculty."""
    from models import get_faculty_attendance_history
    faculty_id = session.get('user_id')
    records = get_faculty_attendance_history(faculty_id, limit=1000)

    history = []
    for r in records:
        total = r['total'] or 0
        present = r['present_count'] or 0
        pct = round((present / total * 100), 1) if total > 0 else 0
        history.append({
            'date': str(r['date']),
            'subject': r['subject'],
            'section': r['section'],
            'department': r.get('department', 'N/A'),
            'period': r['period'],
            'total': total,
            'present': present,
            'absent': r['absent_count'] or 0,
            'percentage': pct,
            'start_time': str(r['start_time']) if r['start_time'] else '--',
            'color': r.get('color', 'transparent')
        })

    return jsonify({'success': True, 'history': history})


@app.route('/api/faculty/update_color', methods=['POST'])
@require_faculty
def api_update_color():
    """Update marking color for a session."""
    data = request.json
    subject = data.get('subject')
    section = data.get('section')
    period = data.get('period')
    date = data.get('date')
    color = data.get('color')

    if not all([subject, section, period, date, color]):
        return jsonify({'success': False, 'error': 'Missing parameters'}), 400

    from models import update_session_color
    if update_session_color(subject, section, period, date, color):
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Update failed'}), 500


@app.route('/api/faculty/delete_session', methods=['POST'])
@require_faculty
def api_delete_session():
    """Delete an attendance session."""
    data = request.json
    subject = data.get('subject')
    section = data.get('section')
    period = data.get('period')
    date = data.get('date')

    if not all([subject, section, period, date]):
        return jsonify({'success': False, 'error': 'Missing parameters'}), 400

    from models import delete_session
    if delete_session(subject, section, period, date):
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Delete failed'}), 500


@app.route('/api/faculty/export_session')
@require_faculty
def api_export_session():
    """Generate Excel for a specific past attendance session."""
    subject = request.args.get('subject', '')
    section = request.args.get('section', '')
    period  = request.args.get('period', '')
    date    = request.args.get('date', '')

    if not subject or not section or not period or not date:
        return jsonify({'error': 'Missing parameters'}), 400

    from models import get_attendance, get_all_students
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    records = get_attendance(subject=subject, section=section, period=period, date=date)
    all_students = get_all_students(section=section)

    # Build lookup of present students
    present_map = {}
    for r in records:
        if r.get('status') == 'PRESENT':
            present_map[r['student_id']] = {
                'time': r.get('time', '--'),
                'confidence': r.get('confidence', 0)
            }

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill  = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = "Attendance Report"; ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Date: {date}"
    ws['A3'] = f"Subject: {subject}"
    ws['A4'] = f"Section: {section}"
    ws['A5'] = f"Period: {period}"
    
    # Get department from the first record if available
    dept = records[0]['department'] if records else session.get('user_dept', 'N/A')
    ws['A6'] = f"Department: {dept}"
    ws['A7'] = f"Faculty: {session.get('user_name', 'N/A')}"

    headers = ['S.No', 'Roll No', 'Student Name', 'Status', 'Time', 'Confidence (%)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin

    for i, student in enumerate(all_students, 1):
        is_present = student['id'] in present_map
        status = 'PRESENT' if is_present else 'ABSENT'
        time_str = present_map[student['id']]['time'] if is_present else '--'
        conf = present_map[student['id']]['confidence'] if is_present else 0
        row_data = [i, student['roll_no'], student['name'], status, time_str,
                    f"{conf:.1f}" if conf else '--']
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=9+i, column=col, value=val)
            cell.border = thin; cell.alignment = Alignment(horizontal='center')
            if col == 4:
                cell.fill = present_fill if is_present else absent_fill

    for col, w in zip(['A','B','C','D','E','F'], [8, 15, 25, 12, 12, 15]):
        ws.column_dimensions[col].width = w

    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    fname = f"Attendance_{subject}_{section}_{date}.xlsx".replace(' ', '_')

    from flask import send_file
    return send_file(buffer,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# ─── Student Image Serving ────────────────────────────────────────────────────
@app.route('/api/student_image/<path:student_dir>')
def get_student_image(student_dir):
    base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dataset')
    full_path = os.path.join(base_dir, student_dir)

    if not os.path.isdir(full_path):
        target = student_dir.lower()
        if os.path.exists(base_dir):
            for d in os.listdir(base_dir):
                if d.lower() == target:
                    full_path = os.path.join(base_dir, d)
                    break

    if os.path.isdir(full_path):
        patterns = ["*.[jJ][pP][gG]", "*.[jJ][pP][eE][gG]", "*.[pP][nN][gG]", "*.[wW][eE][bB][pP]"]
        image_files = []
        for p in patterns:
            image_files.extend(glob.glob(os.path.join(full_path, p)))
        image_files.sort()
        for img_path in image_files:
            if os.path.getsize(img_path) > 0:
                return send_from_directory(os.path.dirname(img_path), os.path.basename(img_path))

    return "Image not found", 404


# ─── Attendance Data API (used by webcam dashboard JS) ───────────────────────
@app.route('/api/attendance_data', methods=['GET'])
def api_attendance_data():
    from attendance_logic import get_session_summary
    from models import get_all_students

    faculty_id = session.get('user_id')
    summary = get_session_summary(faculty_id)
    current_section = summary['session'].get('section') if summary else None
    all_students = get_all_students(section=current_section)

    if not summary:
        absent_list = []
        for s in all_students:
            img_path = s.get('image_path', '')
            folder_name = img_path.replace('dataset/', '').replace('dataset\\', '')
            absent_list.append({
                'id': s['roll_no'], 'name': s['name'],
                'image_url': f"/api/student_image/{folder_name}" if folder_name else None
            })
        return jsonify({'success': True, 'session_active': False, 'present': [],
                        'absent': absent_list, 'total': len(all_students),
                        'present_count': 0, 'absent_count': len(all_students)})

    present_roll_nos = set()
    present = []
    for s in summary.get('present', []):
        present_roll_nos.add(s['roll_no'])
        img_path = s.get('image_path', '')
        folder_name = img_path.replace('dataset/', '').replace('dataset\\', '')
        present.append({'id': s['roll_no'], 'name': s['name'],
                        'time': s.get('time', '--:--'), 'confidence': s.get('confidence', 0),
                        'image_url': f"/api/student_image/{folder_name}" if folder_name else None})

    absent = []
    for s in all_students:
        if s['roll_no'] not in present_roll_nos:
            img_path = s.get('image_path', '')
            folder_name = img_path.replace('dataset/', '').replace('dataset\\', '')
            absent.append({'id': s['roll_no'], 'name': s['name'],
                           'image_url': f"/api/student_image/{folder_name}" if folder_name else None})

    return jsonify({'success': True, 'session_active': True, 'session': summary['session'],
                    'present': present, 'absent': absent, 'total': len(all_students),
                    'present_count': len(present), 'absent_count': len(absent)})


@app.route('/api/sync_students', methods=['POST'])
def api_sync_students():
    from models import sync_enrolled_students, clear_sample_students
    from advanced_face_recognition import get_recognizer
    removed = clear_sample_students()
    synced = sync_enrolled_students()
    # Ensure recognizer sees new students immediately
    get_recognizer().load_students()
    return jsonify({'success': True, 'message': f'Synced {len(synced)} students, removed {removed} old entries', 'synced': synced})


@app.route('/api/students', methods=['GET'])
def api_get_students():
    from models import get_all_students
    students = get_all_students()
    return jsonify({'success': True, 'students': students, 'total': len(students)})


@app.route('/api/reset_session', methods=['POST'])
def api_reset_session():
    from attendance_logic import reset_session
    faculty_id = session.get('user_id')
    success, message = reset_session(faculty_id)
    return jsonify({'success': success, 'message': message})


@app.route('/api/export_excel', methods=['GET'])
def api_export_excel():
    from attendance_logic import get_session_summary
    from models import get_all_students
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    faculty_id = session.get('user_id')
    summary = get_session_summary(faculty_id)
    all_students = get_all_students()
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill  = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    session_info = summary.get('session', {}) if summary else {}
    ws['A1'] = "Attendance Report"; ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Date: {session_info.get('date', datetime.now().strftime('%Y-%m-%d'))}"
    ws['A3'] = f"Subject: {session_info.get('subject', 'N/A')}"
    ws['A4'] = f"Section: {session_info.get('section', 'A')}"
    ws['A5'] = f"Period: {session_info.get('period', 'N/A')}"

    headers = ['S.No', 'Roll No', 'Student Name', 'Status', 'Time', 'Confidence (%)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin

    present_data = {}
    if summary:
        for s in summary.get('present', []):
            present_data[s['roll_no']] = {'time': s.get('time', '--'), 'confidence': s.get('confidence', 0)}

    for i, student in enumerate(all_students, 1):
        roll_no = student['roll_no']
        is_present = roll_no in present_data
        status = 'PRESENT' if is_present else 'ABSENT'
        time_str = present_data[roll_no]['time'] if is_present else '--'
        conf = present_data[roll_no]['confidence'] if is_present else 0
        row_data = [i, roll_no, student['name'], status, time_str, f"{conf:.1f}" if conf else '--']
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=7+i, column=col, value=val)
            cell.border = thin; cell.alignment = Alignment(horizontal='center')
            if col == 4: cell.fill = present_fill if is_present else absent_fill

    for col, w in zip(['A','B','C','D','E','F'], [8, 15, 25, 12, 12, 15]):
        ws.column_dimensions[col].width = w

    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    date_str = session_info.get('date', datetime.now().strftime('%Y-%m-%d'))
    subject  = session_info.get('subject', 'Attendance')

    from flask import send_file
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Attendance_{subject}_{date_str}.xlsx")


# ─── Video Feed ───────────────────────────────────────────────────────────────
@app.route('/video_feed')
def video_feed():
    try:
        from face_recognition_module import WebcamCapture
        from attendance_logic import get_faculty_session, mark_present
        import cv2, time
        faculty_id = session.get('user_id')

        try:
            from advanced_face_recognition import get_recognizer, draw_recognition_boxes
            recognizer = get_recognizer()
            # Ensure students are loaded before starting the feed
            recognizer.load_students()
            use_advanced = True
        except:
            from face_recognition_module import face_manager, draw_face_boxes
            face_manager.load_known_faces()
            use_advanced = False

        def generate_frames():
            cam = WebcamCapture()
            try:
                if not cam.start():
                    yield b'--frame\r\nContent-Type: text/plain\r\n\r\nCamera not available\r\n'
                    return
                frame_count = 0
                last_advanced_results = []
                last_basic_results = []

                while True:
                    frame = cam.read_frame()
                    if frame is None:
                        time.sleep(0.01); continue

                    session_obj = get_faculty_session(faculty_id)
                    display_frame = frame.copy()

                    if session_obj.is_active:
                        try:
                            if use_advanced:
                                # Process every 3rd frame for more responsive voting
                                if frame_count % 3 == 0:
                                    results = recognizer.recognize_frame(frame)
                                    recognizer.add_to_voting_buffer(results)
                                    last_advanced_results = results
                                    voting_result = recognizer.get_voting_result()
                                    if voting_result and voting_result['status'] == 'confirmed':
                                        mark_present(faculty_id, voting_result['name'], voting_result['confidence'] * 100)
                                        # recognizer.clear_voting_buffer() # Keep buffer for continuity or clear if you want 1 ID at a time
                                display_frame = draw_recognition_boxes(display_frame, last_advanced_results)
                            else:
                                if frame_count % 5 == 0:
                                    recognized = face_manager.recognize_faces(frame)
                                    for face in recognized:
                                        if face.get('roll_no') and face['roll_no'] != 'UNKNOWN':
                                            mark_present(faculty_id, face['roll_no'], face['confidence'])
                                    last_basic_results = recognized
                                display_frame = draw_face_boxes(display_frame, last_basic_results)
                        except Exception as e:
                            print(f"Recognition error: {e}")

                    frame_count += 1
                    ret, buf = cv2.imencode('.jpg', display_frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
                    if ret:
                        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf.tobytes() + b'\r\n')
                    if frame_count % 5 != 0:
                        time.sleep(0.01)
            finally:
                cam.stop()
                if use_advanced: recognizer.clear_voting_buffer()

        return app.response_class(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Error Handlers ───────────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e): return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def server_error(e): return jsonify({'error': 'Internal server error'}), 500


# ─── Run ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("\n" + "="*50)
    print("Face Recognition Attendance System")
    print("="*50)
    print("Server starting at: http://localhost:5000")
    print("="*50 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
